# ProductoServicioApp/views.py
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from django.contrib import messages

from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.db.models import Sum, Prefetch
from openpyxl import Workbook, load_workbook

from .models import ProductoExcel, ProductoServicio, Abastecimiento
from .forms import ProductoServicioForm

# 👇 Importamos el modelo de la otra app para prefetchear proveedores con su dirección
from EmpresaPersonaApp.models import EmpresaPersona

IVA_FACTOR = Decimal("1.19")   # 19% IVA
ROUND = lambda x: x.quantize(Decimal("1"), rounding=ROUND_HALF_UP)  # redondeo a entero


def _recalcula_campos(precio_bruto: Decimal, dscto_pct: Decimal | int | None):
    dscto_pct = Decimal(dscto_pct or 0)
    bruto_final = (precio_bruto * (Decimal("1") - dscto_pct / Decimal("100")))
    if bruto_final < 0:
        bruto_final = Decimal("0")

    neto = (bruto_final / IVA_FACTOR)
    iva = bruto_final - neto

    # Redondeo a entero CLP (Decimal entero)
    neto_r = ROUND(neto)
    iva_r = ROUND(iva)
    bruto_r = ROUND(bruto_final)

    # Retornamos ints para asignar directo a PositiveIntegerField
    return int(neto_r), int(iva_r), int(bruto_r)


@transaction.atomic
def lista_productos(request):
    if request.method == 'POST':
        form = ProductoServicioForm(request.POST)
        if form.is_valid():
            # Guardado con cálculo de neto/iva desde bruto y descuento
            producto = form.save(commit=False)

            bruto_in = Decimal(form.cleaned_data.get('produ_bruto') or 0)
            dscto_in = form.cleaned_data.get('produ_dscto') or 0
            neto, iva, bruto_final = _recalcula_campos(bruto_in, dscto_in)

            # Asignar ints a campos Integer del modelo
            producto.produ_neto = neto
            producto.produ_iva = iva
            producto.produ_bruto = bruto_final
            producto.save()

            # Relación con empresa (Abastecimiento)
            empresa = form.cleaned_data['empresa']
            Abastecimiento.objects.get_or_create(emppe=empresa, produ=producto)

            return redirect('productoyservicioapp:lista_productos')
    else:
        form = ProductoServicioForm()

    # ========= CARGA DE PRODUCTOS + PROVEEDOR CON DIRECCIÓN =========
    proveedores_qs = (
        EmpresaPersona.objects
        .select_related(
            'emppe_dire',                 
            'emppe_dire__regi',           
            'emppe_dire__ciuda',          
            'emppe_dire__comun',          
        )
    )


    productos = (
        ProductoServicio.objects
        .prefetch_related(
            Prefetch('proveedores', queryset=proveedores_qs)
        )
        .all()
    )

    # ======= TOTALES =======
    totales = productos.aggregate(
        total_neto=Sum('produ_neto'),
        total_iva=Sum('produ_iva'),
        total_bruto=Sum('produ_bruto'),
    )

    # 👇👇 AQUÍ armamos el diccionario para el JS
    adjuntos_map: dict[str, list[dict]] = {}

    for producto in productos:
        adjuntos = []
        for adj in producto.archivos_excel.all():
            extension = Path(adj.archivo.name).suffix.replace(".", "").upper()
            adjuntos.append(
                {
                    "id": adj.pk,
                    "sku": adj.producto.produ_sku,
                    "nombre": adj.nombre_archivo,
                    "url": adj.archivo.url if adj.archivo else "",
                    "neto": adj.valor_neto,
                    "iva": adj.valor_iva,
                    "bruto": adj.valor_bruto,
                    "descuento": adj.descuento,
                    "tipo": extension or "ARCHIVO",
                }
            )

        producto.archivos_detalle = adjuntos
        producto.archivos_count = len(adjuntos)

        adjuntos_map[producto.pk] = adjuntos

    context = {
        'productos': productos,
        'form': form,
        'total_neto':  totales.get('total_neto')  or 0,
        'total_iva':   totales.get('total_iva')   or 0,
        'total_bruto': totales.get('total_bruto') or 0,
        'adjuntos_map': adjuntos_map,
    }
    return render(request, 'producto_servicio/producto_servicio.html', context)


@transaction.atomic
def editar_producto(request, pk):
    producto = get_object_or_404(ProductoServicio, pk=pk)
    abastecimiento = Abastecimiento.objects.filter(produ=producto).first()
    empresa_actual = abastecimiento.emppe if abastecimiento else None

    if request.method == 'POST':
        form = ProductoServicioForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save(commit=False)

            bruto_in = Decimal(form.cleaned_data.get('produ_bruto') or 0)
            dscto_in = form.cleaned_data.get('produ_dscto') or 0
            neto, iva, bruto_final = _recalcula_campos(bruto_in, dscto_in)

            producto.produ_neto = neto
            producto.produ_iva = iva
            producto.produ_bruto = bruto_final
            producto.save()

            # Empresa/proveedor
            nueva_empresa = form.cleaned_data['empresa']
            if empresa_actual != nueva_empresa:
                if empresa_actual:
                    Abastecimiento.objects.filter(emppe=empresa_actual, produ=producto).delete()
                Abastecimiento.objects.get_or_create(emppe=nueva_empresa, produ=producto)

            return redirect('productoyservicioapp:lista_productos')
    else:
        initial_data = {'empresa': empresa_actual} if empresa_actual else {}
        form = ProductoServicioForm(instance=producto, initial=initial_data)

    return render(request, 'producto_servicio/producto_servicio.html', {'form': form, 'producto': producto})


@transaction.atomic
def eliminar_producto(request, pk):
    producto = get_object_or_404(ProductoServicio, pk=pk)
    if request.method == 'POST':
        Abastecimiento.objects.filter(produ=producto).delete()
        producto.delete()
    return redirect('productoyservicioapp:lista_productos')

def descargar_plantilla_productos(request):
    # Crear libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos o Servicios"

    # Encabezados de la tabla 
    headers = [
        "PRODU_NOM",
        "PRODU_DESC",
        "PRODU_BRUTO",
        "PRODU_DSCTO",
        "EMPPE_ID (proveedor)",
    ]

    # Agregar encabezados en la primera fila
    ws.append(headers)

    # Fila de ejemplo para guiar el llenado
    ws.append([
        "Nombre producto o servicio",
        "Descripción o detalles",
        10000,
        0,
        None,  # Dejar en blanco u optar por un ID numérico de proveedor existente
    ])

    # Preparar respuesta HTTP para descarga
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="plantilla_productos_servicios.xlsx"'
    )

    wb.save(response)
    return response


@transaction.atomic
def cargar_productos_excel(request):
    if request.method != "POST":
        messages.error(request, "Debes seleccionar un archivo Excel para importar.")
        return redirect("productoyservicioapp:lista_productos")

    archivo = request.FILES.get("archivo_excel")
    if not archivo:
        messages.error(request, "Selecciona un archivo .xlsx con la plantilla de productos.")
        return redirect("productoyservicioapp:lista_productos")

    try:
        workbook = load_workbook(archivo)
    except Exception:
        messages.error(request, "No se pudo leer el archivo. Verifica que sea un .xlsx válido.")
        return redirect("productoyservicioapp:lista_productos")

    sheet = workbook.active
    expected_headers = ["PRODU_NOM", "PRODU_DESC", "PRODU_BRUTO", "PRODU_DSCTO", "EMPPE_ID (proveedor)"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    if headers[: len(expected_headers)] != expected_headers:
        messages.error(
            request,
            "La plantilla no coincide con el formato esperado. Descarga nuevamente la plantilla y vuelve a intentarlo.",
        )
        return redirect("productoyservicioapp:lista_productos")

    creados = 0
    errores: list[str] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        nombre, descripcion, bruto, dscto, proveedor_id = row[:5]

        if not any(row):
            continue  # fila completamente vacía

        if not nombre or not descripcion or bruto is None:
            errores.append(f"Fila {row_number}: nombre, descripción y valor bruto son obligatorios.")
            continue

        try:
            bruto_decimal = Decimal(bruto)
        except Exception:
            errores.append(f"Fila {row_number}: el valor bruto debe ser numérico.")
            continue

        dscto_value = dscto or 0
        try:
            dscto_decimal = Decimal(dscto_value)
        except Exception:
            errores.append(f"Fila {row_number}: el descuento debe ser un número entre 0 y 100.")
            continue

        if dscto_decimal < 0 or dscto_decimal > 100:
            errores.append(f"Fila {row_number}: el descuento debe estar entre 0 y 100%.")
            continue

        neto, iva, bruto_final = _recalcula_campos(bruto_decimal, dscto_decimal)

        producto = ProductoServicio(
            produ_nom=str(nombre).strip(),
            produ_desc=str(descripcion).strip(),
            produ_bruto=bruto_final,
            produ_neto=neto,
            produ_iva=iva,
            produ_dscto=int(dscto_decimal),
        )
        producto.save()
        creados += 1

        if proveedor_id:
            try:
                proveedor_id_int = int(proveedor_id)
            except (TypeError, ValueError):
                errores.append(
                    f"Fila {row_number}: el ID de proveedor debe ser numérico; el producto se creó sin proveedor."
                )
                continue

            proveedor = EmpresaPersona.objects.filter(pk=proveedor_id_int).first()
            if proveedor:
                Abastecimiento.objects.get_or_create(emppe=proveedor, produ=producto)
            else:
                errores.append(
                    f"Fila {row_number}: no existe proveedor con ID {proveedor_id_int}; el producto se creó sin proveedor."
                )

    if creados:
        messages.success(request, f"Se importaron {creados} productos/servicios desde el Excel.")

    if errores:
        for err in errores:
            messages.warning(request, err)

    return redirect("productoyservicioapp:lista_productos")


def _procesar_excel_producto(archivo) -> tuple[int, int, int, int]:
    try:
        workbook = load_workbook(archivo)
    except Exception:
        raise ValueError("No se pudo leer el archivo. Verifica que sea un .xlsx válido.")

    sheet = workbook.active
    expected_headers = ["PRODU_NOM", "PRODU_DESC", "PRODU_BRUTO", "PRODU_DSCTO", "EMPPE_ID (proveedor)"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    if headers[: len(expected_headers)] != expected_headers:
        raise ValueError(
            "La plantilla no coincide con el formato esperado. Descarga nuevamente la plantilla y vuelve a intentarlo."
        )

    primera_fila = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            primera_fila = row
            break

    if not primera_fila:
        raise ValueError("El archivo no contiene filas de datos.")

    _, _, bruto, dscto, _ = primera_fila[:5]

    if bruto is None:
        raise ValueError("La fila de datos debe incluir un valor bruto.")

    try:
        bruto_decimal = Decimal(bruto)
    except Exception:
        raise ValueError("El valor bruto debe ser numérico.")

    dscto_value = dscto or 0
    try:
        dscto_decimal = Decimal(dscto_value)
    except Exception:
        raise ValueError("El descuento debe ser un número entre 0 y 100.")

    if dscto_decimal < 0 or dscto_decimal > 100:
        raise ValueError("El descuento debe estar entre 0 y 100%.")

    neto, iva, bruto_final = _recalcula_campos(bruto_decimal, dscto_decimal)
    return neto, iva, bruto_final, int(dscto_decimal)


@transaction.atomic
def subir_excel_producto(request, pk):
    producto = get_object_or_404(ProductoServicio, pk=pk)

    if request.method != "POST":
        messages.error(request, "Debes seleccionar un archivo para este producto.")
        return redirect("productoyservicioapp:lista_productos")

    archivos = request.FILES.getlist("archivos_excel")
    if not archivos:
        messages.error(request, "Selecciona al menos un archivo.")
        return redirect("productoyservicioapp:lista_productos")

    procesados = 0
    errores: list[str] = []
    actualizados = 0

    for archivo in archivos:
        extension = Path(archivo.name).suffix.lower()
        es_excel = extension in {".xlsx", ".xls", ".xlsm"}

        if es_excel:
            try:
                archivo.seek(0)
                neto, iva, bruto_final, dscto = _procesar_excel_producto(archivo)
                archivo.seek(0)
            except ValueError as exc:
                errores.append(f"{archivo.name}: {exc}")
                continue

            producto.produ_neto = neto
            producto.produ_iva = iva
            producto.produ_bruto = bruto_final
            producto.produ_dscto = dscto
            producto.save(update_fields=["produ_neto", "produ_iva", "produ_bruto", "produ_dscto"])
            actualizados += 1
        else:
            neto = producto.produ_neto
            iva = producto.produ_iva
            bruto_final = producto.produ_bruto
            dscto = producto.produ_dscto or 0

        ProductoExcel.objects.create(
            producto=producto,
            archivo=archivo,
            valor_neto=neto,
            valor_iva=iva,
            valor_bruto=bruto_final,
            descuento=dscto,
        )

        procesados += 1

    if procesados:
        detalle_actualizados = (
            f" y se actualizaron sus datos contables con {actualizados} archivo(s) Excel"
            if actualizados
            else ""
        )
        messages.success(
            request,
            f"Se cargaron {procesados} archivo(s) para '{producto.produ_nom}'{detalle_actualizados}.",
        )

    if errores:
        for err in errores:
            messages.warning(request, err)

    return redirect("productoyservicioapp:lista_productos")

@transaction.atomic
def eliminar_archivos_producto(request, pk):
    producto = get_object_or_404(ProductoServicio, pk=pk)

    if request.method != "POST":
        messages.error(request, "Solicitud inválida para eliminación de archivos.")
        return redirect("productoyservicioapp:lista_productos")

    ids_str = request.POST.get("archivos_ids", "")
    ids = [int(x) for x in ids_str.split(",") if x.strip().isdigit()]

    if not ids:
        messages.warning(request, "No se seleccionó ningún archivo para eliminar.")
        return redirect("productoyservicioapp:lista_productos")

    eliminados, _ = ProductoExcel.objects.filter(
        producto=producto,
        pk__in=ids
    ).delete()

    if eliminados:
        messages.success(
            request,
            f"Se eliminaron {eliminados} archivo(s) asociado(s) al producto '{producto.produ_nom}'."
        )
    else:
        messages.warning(
            request,
            "No se encontraron archivos para eliminar. Intenta nuevamente."
        )

    return redirect("productoyservicioapp:lista_productos")